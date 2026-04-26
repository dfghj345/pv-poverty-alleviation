from __future__ import annotations

import argparse
import asyncio
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import hashlib
from pathlib import Path
import re
import sys
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import delete, func, select
from sqlalchemy.dialects.postgresql import insert

PROJECT_ROOT = Path(__file__).resolve().parents[2]
BACKEND_DIR = Path(__file__).resolve().parents[1]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from app.models.panel_data import COLUMN_MAPPING, COMPARISON_COLUMNS, PanelData, UPSERT_COLUMNS

DEFAULT_FILE = PROJECT_ROOT / "data" / "raw" / "面板数据.xlsx"
REPORT_SHEET_NAME = "报告"
NULL_LITERALS = {"", "-", "--", "—", "N/A", "NA", "null", "None", "无", "暂无"}
RAW_EXCEL_COLUMNS = tuple(key for key in COLUMN_MAPPING.keys() if key != "省份")
NUMERIC_FIELDS = {
    "pv_installed_capacity_wan_kw",
    "disposable_income_per_capita_yuan",
    "healthcare_expenditure_per_capita_yuan",
    "urban_rural_income_ratio",
    "mortality_per_mille",
    "pm25_annual_avg_ug_per_m3",
    "gdp_100m_yuan",
}
PERCENT_FIELDS: set[str] = set()
HEADER_ALIASES = {
    "城市": "city",
    "年份": "year",
    "pv(光伏装机量,万千瓦)": "pv_installed_capacity_wan_kw",
    "income(人均可支配收入,元)": "disposable_income_per_capita_yuan",
    "health(人均医疗支出,元)": "healthcare_expenditure_per_capita_yuan",
    "ur_ratio(城乡收入比)": "urban_rural_income_ratio",
    "mortality(总死亡率,‰)": "mortality_per_mille",
    "pm25(年均浓度,μg/m³)": "pm25_annual_avg_ug_per_m3",
    "gdp(地区生产总值,亿元)": "gdp_100m_yuan",
}
CHINESE_DIGIT_MAP = {
    "零": 0,
    "〇": 0,
    "一": 1,
    "二": 2,
    "两": 2,
    "三": 3,
    "四": 4,
    "五": 5,
    "六": 6,
    "七": 7,
    "八": 8,
    "九": 9,
}
CHINESE_UNIT_MAP = {
    "十": 10,
    "百": 100,
    "千": 1000,
    "万": 10000,
    "亿": 100000000,
}


@dataclass(slots=True)
class SheetSource:
    sheet_name: str
    province: str
    workbook_name: str | None


@dataclass(slots=True)
class FailureRecord:
    sheet: str
    row_number: int
    reason: str


@dataclass(slots=True)
class ImportStats:
    total_rows: int = 0
    inserted: int = 0
    updated: int = 0
    skipped: int = 0
    failed: int = 0
    failures: list[FailureRecord] = field(default_factory=list)

    def add_failure(self, *, sheet: str, row_number: int, reason: str) -> None:
        self.failed += 1
        if len(self.failures) < 5:
            self.failures.append(FailureRecord(sheet=sheet, row_number=row_number, reason=reason))


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text in NULL_LITERALS:
        return None
    return text


def normalize_header(value: Any) -> str:
    text = normalize_text(value)
    if text is None:
        return ""
    text = text.replace("（", "(").replace("）", ")")
    text = text.replace("﹪", "%").replace("％", "%")
    text = re.sub(r"\s+", "", text)
    return text.lower()


def chinese_to_decimal(text: str) -> Decimal:
    if not text:
        raise ValueError("空中文数字")
    if "点" in text:
        integer_part, decimal_part = text.split("点", 1)
        base = chinese_to_decimal(integer_part) if integer_part else Decimal("0")
        decimals = "".join(str(CHINESE_DIGIT_MAP[ch]) for ch in decimal_part if ch in CHINESE_DIGIT_MAP)
        return Decimal(f"{base}.{decimals}") if decimals else base

    total = 0
    section = 0
    number = 0
    for ch in text:
        if ch in CHINESE_DIGIT_MAP:
            number = CHINESE_DIGIT_MAP[ch]
            continue
        unit = CHINESE_UNIT_MAP.get(ch)
        if unit is None:
            raise ValueError(f"无法识别中文数字: {text}")
        if unit < 10000:
            section += (number or 1) * unit
        else:
            section = (section + number) * unit
            total += section
            section = 0
        number = 0
    return Decimal(total + section + number)


def normalize_decimal(value: Any, *, is_percent: bool = False) -> Decimal | None:
    if value is None:
        return None

    if isinstance(value, Decimal):
        number = value
    elif isinstance(value, bool):
        return None
    elif isinstance(value, (int, float)):
        number = Decimal(str(value))
    else:
        text = normalize_text(value)
        if text is None:
            return None
        text = text.replace(",", "").replace("，", "").replace(" ", "")
        percent = text.endswith("%") or text.endswith("％")
        if percent:
            text = text[:-1]
        unit_multiplier = Decimal("1")
        if text.endswith("万") and text != "万":
            unit_multiplier = Decimal("10000")
            text = text[:-1]
        elif text.endswith("亿") and text != "亿":
            unit_multiplier = Decimal("100000000")
            text = text[:-1]
        if any(ch in CHINESE_DIGIT_MAP or ch in CHINESE_UNIT_MAP or ch == "点" for ch in text):
            number = chinese_to_decimal(text)
        else:
            try:
                number = Decimal(text)
            except InvalidOperation as exc:
                raise ValueError(f"无法解析数值: {value}") from exc
        number *= unit_multiplier
        if percent or is_percent:
            number = number / Decimal("100")
        return number

    if is_percent:
        number = number / Decimal("100")
    return number


def normalize_year(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.year
    if isinstance(value, date):
        return value.year
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return int(value)
    text = normalize_text(value)
    if text is None:
        return None
    match = re.search(r"(19|20)\d{2}", text)
    if match:
        return int(match.group(0))
    number = normalize_decimal(text)
    return int(number) if number is not None else None


def normalize_date(value: Any) -> date | datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return value
    text = normalize_text(value)
    if text is None:
        return None
    text = text.replace(".", "-").replace("/", "-").replace("年", "-").replace("月", "-").replace("日", "")
    for fmt in ("%Y-%m-%d", "%Y-%m", "%Y%m%d"):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.date()
        except ValueError:
            continue
    raise ValueError(f"无法解析日期: {value}")


def build_data_hash(province: str, city: str, year: int) -> str:
    payload = f"{province}|{city}|{year}"
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def detect_header_row(ws: Worksheet) -> int:
    best_row = 1
    best_count = -1
    for row_number in range(1, min(ws.max_row, 10) + 1):
        values = [normalize_text(ws.cell(row_number, col).value) for col in range(1, ws.max_column + 1)]
        count = sum(1 for value in values if value)
        if count > best_count:
            best_count = count
            best_row = row_number
    return best_row


def build_merged_lookup(ws: Worksheet) -> dict[tuple[int, int], Any]:
    merged_lookup: dict[tuple[int, int], Any] = {}
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_value = ws.cell(min_row, min_col).value
        for row_number in range(min_row, max_row + 1):
            for col_number in range(min_col, max_col + 1):
                merged_lookup[(row_number, col_number)] = top_left_value
    return merged_lookup


def read_cell(ws: Worksheet, row_number: int, col_number: int, merged_lookup: dict[tuple[int, int], Any]) -> Any:
    value = ws.cell(row_number, col_number).value
    if value is not None:
        return value
    return merged_lookup.get((row_number, col_number))


def resolve_sheet_sources(workbook) -> list[SheetSource]:
    sources: list[SheetSource] = []
    if REPORT_SHEET_NAME in workbook.sheetnames:
        report = workbook[REPORT_SHEET_NAME]
        for row_number in range(5, report.max_row + 1):
            workbook_name = normalize_text(report.cell(row_number, 2).value)
            sheet_name = normalize_text(report.cell(row_number, 5).value)
            if not workbook_name or not sheet_name or workbook_name.startswith("工作簿"):
                continue
            normalized_sheet_name = sheet_name.strip("'")
            if normalized_sheet_name not in workbook.sheetnames:
                continue
            province = workbook_name.split("_", 1)[0].strip()
            sources.append(
                SheetSource(
                    sheet_name=normalized_sheet_name,
                    province=province,
                    workbook_name=workbook_name,
                )
            )
    if sources:
        return sources

    for sheet_name in workbook.sheetnames:
        if sheet_name == REPORT_SHEET_NAME:
            continue
        sources.append(SheetSource(sheet_name=sheet_name, province="", workbook_name=None))
    return sources


def resolve_header_mapping(ws: Worksheet, header_row: int) -> dict[int, str]:
    column_map: dict[int, str] = {}
    for col_number in range(1, ws.max_column + 1):
        raw_header = ws.cell(header_row, col_number).value
        normalized_header = normalize_header(raw_header)
        field_name = HEADER_ALIASES.get(normalized_header)
        if field_name:
            column_map[col_number] = field_name

    missing = [header for header in RAW_EXCEL_COLUMNS if COLUMN_MAPPING[header] not in column_map.values()]
    if missing:
        missing_text = ", ".join(missing)
        raise ValueError(f"缺少表头: {missing_text}")
    return column_map


def normalize_row(
    *,
    raw_row: dict[str, Any],
    province: str,
    file_path: Path,
    workbook_name: str | None,
    sheet_name: str,
    row_number: int,
) -> dict[str, Any]:
    city = normalize_text(raw_row.get("city"))
    year = normalize_year(raw_row.get("year"))
    if not province:
        raise ValueError("缺少省份信息")
    if not city:
        raise ValueError("缺少城市")
    if year is None:
        raise ValueError("缺少年份")

    record: dict[str, Any] = {
        "province": province,
        "city": city,
        "year": year,
        "source_file": str(file_path),
        "source_workbook": workbook_name,
        "source_sheet": sheet_name,
        "source_row_number": row_number,
    }

    for field_name in NUMERIC_FIELDS:
        record[field_name] = normalize_decimal(raw_row.get(field_name), is_percent=field_name in PERCENT_FIELDS)

    record["data_hash"] = build_data_hash(province=province, city=city, year=year)
    return record


def records_equal(left: dict[str, Any], right: dict[str, Any]) -> bool:
    for field_name in COMPARISON_COLUMNS:
        if left.get(field_name) != right.get(field_name):
            return False
    return True


def collect_records(file_path: Path) -> tuple[list[dict[str, Any]], ImportStats]:
    workbook = load_workbook(file_path, read_only=False, data_only=True)
    stats = ImportStats()
    deduped_records: dict[str, dict[str, Any]] = {}
    sources = resolve_sheet_sources(workbook)

    for source in sources:
        ws = workbook[source.sheet_name]
        header_row = detect_header_row(ws)
        merged_lookup = build_merged_lookup(ws)
        try:
            header_mapping = resolve_header_mapping(ws, header_row)
        except ValueError as exc:
            stats.add_failure(sheet=source.sheet_name, row_number=header_row, reason=str(exc))
            continue

        previous_city: Any = None
        for row_number in range(header_row + 1, ws.max_row + 1):
            raw_row: dict[str, Any] = {}
            has_value = False
            for col_number, field_name in header_mapping.items():
                value = read_cell(ws, row_number, col_number, merged_lookup)
                if field_name == "city" and value is None:
                    value = previous_city
                if value is not None and str(value).strip() != "":
                    has_value = True
                raw_row[field_name] = value

            if not has_value:
                stats.skipped += 1
                continue

            stats.total_rows += 1
            if normalize_text(raw_row.get("city")) is not None:
                previous_city = raw_row.get("city")

            try:
                normalized = normalize_row(
                    raw_row=raw_row,
                    province=source.province,
                    file_path=file_path,
                    workbook_name=source.workbook_name,
                    sheet_name=source.sheet_name,
                    row_number=row_number,
                )
            except ValueError as exc:
                stats.add_failure(sheet=source.sheet_name, row_number=row_number, reason=str(exc))
                continue

            existing = deduped_records.get(normalized["data_hash"])
            if existing is None:
                deduped_records[normalized["data_hash"]] = normalized
                continue
            if records_equal(existing, normalized):
                stats.skipped += 1
                continue
            deduped_records[normalized["data_hash"]] = normalized
            stats.skipped += 1

    return list(deduped_records.values()), stats


def chunked(items: list[dict[str, Any]], chunk_size: int) -> list[list[dict[str, Any]]]:
    return [items[index:index + chunk_size] for index in range(0, len(items), chunk_size)]


async def fetch_existing_records(hashes: list[str], *, allow_failure: bool) -> dict[str, dict[str, Any]]:
    if not hashes:
        return {}

    try:
        from app.db.session import AsyncSessionLocal
    except Exception:
        if allow_failure:
            return {}
        raise

    existing: dict[str, dict[str, Any]] = {}
    try:
        async with AsyncSessionLocal() as session:
            for current_chunk in [hashes[index:index + 1000] for index in range(0, len(hashes), 1000)]:
                result = await session.execute(select(PanelData).where(PanelData.data_hash.in_(current_chunk)))
                for model in result.scalars():
                    existing[model.data_hash] = {
                        "data_hash": model.data_hash,
                        "province": model.province,
                        "city": model.city,
                        "year": model.year,
                        "pv_installed_capacity_wan_kw": model.pv_installed_capacity_wan_kw,
                        "disposable_income_per_capita_yuan": model.disposable_income_per_capita_yuan,
                        "healthcare_expenditure_per_capita_yuan": model.healthcare_expenditure_per_capita_yuan,
                        "urban_rural_income_ratio": model.urban_rural_income_ratio,
                        "mortality_per_mille": model.mortality_per_mille,
                        "pm25_annual_avg_ug_per_m3": model.pm25_annual_avg_ug_per_m3,
                        "gdp_100m_yuan": model.gdp_100m_yuan,
                        "source_file": model.source_file,
                        "source_workbook": model.source_workbook,
                        "source_sheet": model.source_sheet,
                        "source_row_number": model.source_row_number,
                    }
    except Exception:
        if allow_failure:
            return {}
        raise

    return existing


def split_changes(
    records: list[dict[str, Any]],
    existing_records: dict[str, dict[str, Any]],
    stats: ImportStats,
) -> list[dict[str, Any]]:
    rows_to_write: list[dict[str, Any]] = []
    for record in records:
        existing = existing_records.get(record["data_hash"])
        if existing is None:
            stats.inserted += 1
            rows_to_write.append(record)
            continue
        if records_equal(existing, record):
            stats.skipped += 1
            continue
        stats.updated += 1
        rows_to_write.append(record)
    return rows_to_write


async def write_records(
    rows_to_write: list[dict[str, Any]],
    *,
    truncate: bool,
    batch_size: int,
) -> None:
    from app.db.session import AsyncSessionLocal

    async with AsyncSessionLocal() as session:
        try:
            if truncate:
                await session.execute(delete(PanelData))
            for batch in chunked(rows_to_write, batch_size):
                stmt = insert(PanelData).values(batch)
                stmt = stmt.on_conflict_do_update(
                    index_elements=["data_hash"],
                    set_={
                        **{column: getattr(stmt.excluded, column) for column in UPSERT_COLUMNS},
                        "updated_at": func.now(),
                    },
                )
                await session.execute(stmt)
            await session.commit()
        except Exception:
            await session.rollback()
            raise


async def ensure_panel_data_table() -> None:
    from app.db.session import engine, initialize_database

    await initialize_database()
    async with engine.begin() as conn:
        await conn.run_sync(PanelData.__table__.create, checkfirst=True)


async def run_import(args: argparse.Namespace) -> int:
    file_path = Path(args.file).expanduser().resolve()
    if not file_path.exists():
        print(f"文件不存在: {file_path}")
        return 1

    records, stats = collect_records(file_path)
    existing_records: dict[str, dict[str, Any]] = {}
    rows_to_write: list[dict[str, Any]]

    if not args.dry_run:
        await ensure_panel_data_table()

    if args.truncate:
        stats.inserted = len(records)
        rows_to_write = records
    else:
        existing_records = await fetch_existing_records(
            [record["data_hash"] for record in records],
            allow_failure=bool(args.dry_run),
        )
        rows_to_write = split_changes(records, existing_records, stats)

    if not args.dry_run and rows_to_write:
        await write_records(rows_to_write, truncate=bool(args.truncate), batch_size=int(args.batch_size))

    print(f"读取总行数: {stats.total_rows}")
    print(f"成功导入数: {stats.inserted}")
    print(f"更新数: {stats.updated}")
    print(f"跳过数: {stats.skipped}")
    print(f"失败数: {stats.failed}")
    if stats.failures:
        print("前 5 条失败原因:")
        for failure in stats.failures:
            print(f"- {failure.sheet} 第 {failure.row_number} 行: {failure.reason}")
    return 0 if stats.failed == 0 else 2


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import panel Excel data into PostgreSQL.")
    parser.add_argument("--file", default=str(DEFAULT_FILE), help="Excel file path.")
    parser.add_argument("--dry-run", action="store_true", help="Validate and compare only, do not write.")
    parser.add_argument("--truncate", action="store_true", help="Delete existing panel_data rows before import.")
    parser.add_argument("--batch-size", type=int, default=500, help="Batch size for database upsert.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    return asyncio.run(run_import(args))


if __name__ == "__main__":
    raise SystemExit(main())
